from bs4 import BeautifulSoup as bs
from pathlib import Path
from typing import Optional,Union,Dict,List
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import os
import re
import requests as rq
import json
from multiprocessing import Pool
import numpy as np


def get_headers(
    key: str,
    default_value: Optional[str] = None
    )-> Dict[str,Dict[str,str]]:
    """ Get Headers """
    JSON_FILE : str = 'Coupang_Cawling/json/headers.json'

    with open(JSON_FILE,'r',encoding='UTF-8') as file:
        headers : Dict[str,Dict[str,str]] = json.loads(file.read())

    try :
        return headers[key]
    except:
        if default_value:
            return default_value
        raise EnvironmentError(f'Set the {key}')
    


class Coupang:

    def __init__(self)-> None:
        self.__headers : Dict[str,str] = get_headers(key='headers')

    def get_product_code(self, product_link_list: list)-> list:
        """ 입력받은 URL 주소의 PRODUCT CODE 추출하는 메소드 """
        prod_code_list = list()
        for prod_link in product_link_list:
            prod_code : str = prod_link.split('itemId=')[1].split('&')[0]
            prod_code_list.append(prod_code)
        return prod_code_list

    def get_product_links(self, product_list_link: str)-> list:
        """상품 목록 링크를 입력하면 상품들의 링크 추출하여 리스트로 반환하는 함수"""
        headers = self.__headers # 헤더 불러오기
        links = list() # 반환하는 링크 리스트 생성
        review_counts = list() # 리뷰 수 리스트 생성

        with rq.Session() as session:
            response = session.get(product_list_link, headers=headers) # 불러온 헤더로 리퀘스트
            soup = bs(response.text, 'html.parser') # 객체 생성
            li_tags = soup.select('div > section > form > div:nth-of-type(2) > div:nth-of-type(2) > ul > li') # 상품 id가 저장된 태그 전부 크롤링
            for li in li_tags: # 링크 재가공
                id = li.get('id') # 태그에서 id추출
                vendor_item_id = li.get('data-vendor-item-id') # 태그에서 vendor_item_id 추출
                link = f'https://www.coupang.com/vp/products/5999714344?itemId={id}&vendorItemId={vendor_item_id}&pickType=COU_PICK&sourceType=srp_product_ads&clickEventId=67ce6baa-73b3-4b93-8325-9476b7933790&korePlacement=15&koreSubPlacement=1&clickEventId=67ce6baa-73b3-4b93-8325-9476b7933790&korePlacement=15&koreSubPlacement=1&q=%ED%82%A4%EB%B3%B4%EB%93%9C&itemsCount=36&searchId=f3ae5c0ca337413a95aad71cd823b793&rank=0'
                # 링크 재가공 코드
                links.append(link) # 링크 리스트에 추가

            # 리뷰 수 추출
            review_count_tags = soup.select('span.rating-total-count')
            for review_count_tag in review_count_tags:
                review_count = int(re.findall(r'\d+', review_count_tag.text)[0])
                if review_count//5 > 140:
                    review_counts.append(140)
                else:
                    review_counts.append(review_count//5+1)

            return links, review_counts

    def main(self, product_list_links):
        """모든 함수들 모아서 크롤링 진행하는 main 함수"""
        all_results = []
        for product_list_link in product_list_links:
            # URL 주소
            product_links, review_counts = self.get_product_links(product_list_link=product_list_link)
            # URL의 Product Code 추출
            prod_codes : str = self.get_product_code(product_link_list=product_links)

            # URL 주소 재가공
            URLS : List[str] = []
            for prod_code, page_count in zip(prod_codes, review_counts):
                for page in range(1, page_count):
                    url = f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={page}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true'
                    URLS.append(url)

            # __headers에 referer 키 추가
            self.__headers['referer'] = product_list_link

            with rq.Session() as session:
                with Pool(processes=5) as pool:
                    results = pool.starmap(self.fetch, [(url, session) for url in URLS])
                    all_results.extend(results)

        return all_results
    
    def fetch(self, url:str,session)-> List[Dict[str,Union[str,int]]]:
        """실질적인 크롤링 데이터 생성 함수"""
        save_data : List[Dict[str,Union[str,int]]] = list()
        prod_code : str = url.split('productId=')[1].split('&')[0]
        
        with session.get(url=url,headers=self.__headers) as response :
            html = response.text
            soup = bs(html,'html.parser')

            # Article Boxes
            article_lenth = len(soup.select('article.sdp-review__article__list'))

            for idx in range(article_lenth):
                dict_data : Dict[str,Union[str,int]] = dict()
                articles = soup.select('article.sdp-review__article__list')

                # 평점
                rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange.js_reviewArticleRatingValue')
                if rating == None:
                    rating = 0
                else :
                    rating = int(rating.attrs['data-rating'])

                # 구매자 상품명
                prod_name = articles[idx].select_one('div.sdp-review__article__list__info__product-info__name')
                if prod_name == None or prod_name.text == '':
                    prod_name = '-'
                else:
                    prod_name = prod_name.text.strip()

                # 헤드라인(타이틀)
                headline = articles[idx].select_one('div.sdp-review__article__list__headline')
                if headline == None or headline.text == '':
                    headline = '등록된 헤드라인이 없습니다'
                else:
                    headline = headline.text.strip()

                # 리뷰 내용
                review_content = articles[idx].select_one('div.sdp-review__article__list__review__content.js_reviewArticleContent')
                if review_content == None :
                    review_content = '등록된 리뷰내용이 없습니다'
                else:
                    review_content = re.sub('[\n\t]','',review_content.text.strip())

                dict_data['prod_name'] = prod_name
                dict_data['prod_number'] = prod_code
                dict_data['rating'] = rating
                dict_data['headline'] = headline
                dict_data['review_content'] = review_content
                

                save_data.append(dict_data)

                print(dict_data , '\n')

            time.sleep(np.random.randint(1, 2))

            return save_data

    def input_review_url(self)-> str:
        """상품 link 입력 함수"""
        while True:
            # Window
            # os.system('cls')
            # Mac
            os.system('clear')
            
            # Review URL
            review_url : str = input('원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906&q=%ED%9E%98%EB%82%B4%EB%B0%94+%EC%B4%88%EC%BD%94+%EC%8A%A4%EB%8B%88%EC%BB%A4%EC%A6%88&itemsCount=36&searchId=0c5c84d537bc41d1885266961d853179&rank=2&isAddedCart=\n\n:')
            if not review_url :
                # Window
                os.system('cls')
                # Mac
                #os.system('clear')
                print('URL 주소가 입력되지 않았습니다')
                continue
            return review_url

    def input_page_count(self)-> int:
        """리뷰 페이지 수 입력 함수"""
        # Window
        # os.system('cls')
        # Mac
        os.system('clear')
        while True:
            page_count : str = input('페이지 수를 입력하세요\n\n:')
            if not page_count:
                print('페이지 수가 입력되지 않았습니다\n')
                continue

            return int(page_count)

def clean_string(s):
    """
    Remove illegal characters from the string for Excel compatibility
    """
    return ''.join(c for c in s if c == '\n' or c == '\t' or 0x20 <= ord(c) < 0xD7FF or 0xE000 <= ord(c) < 0xFFFE or 0x10000 <= ord(c) < 0x10FFFF)

base_url = "https://www.coupang.com/np/search?q=%EC%8B%9C%EB%8B%88%EC%96%B4+%EA%B1%B4%EA%B0%95%EC%8B%9D%ED%92%88&channel=user&component=&eventCategory=SRP&trcid=&traid=&sorter=scoreDesc&minPrice=&maxPrice=&priceRange=&filterType=&listSize=36&filter=&isPriceRange=false&brand=&offerCondition=&rating=0&page={}&rocketAll=false&searchIndexingToken=1=9&backgroundColor="
            
urls = [base_url.format(page) for page in range(1, 21)]

class OpenPyXLNo:
    @staticmethod
    def save_file()-> None:
        """결과 저장 함수"""
        # 크롤링 결과
        results : List[List[Dict[str,Union[str,int]]]] = Coupang().main(product_list_links=urls)

        savePath : str = os.path.abspath('쿠팡-상품리뷰-크롤링')
        fileName : str = 'reviwe_헬스건강식품_시니어 건강식품' + '.xlsx'
        filePath : str = os.path.join(savePath,fileName)

        # 파일이 존재하는지 확인
        if os.path.exists(filePath):
            # 파일이 존재하면, 파일을 불러옵니다.
            wb = load_workbook(filePath)
            ws = wb.active
        else:
            # 파일이 존재하지 않으면, 새로운 파일을 생성합니다.
            wb = Workbook()
            ws = wb.active
            ws.append(['상품명', '상품 코드', '구매자 평점', '리뷰 제목', '리뷰 내용'])

        # 마지막 행 번호를 찾습니다.
        row = ws.max_row + 1

        for x in results:
            for result in x:
                ws[f'A{row}'] = clean_string(result['prod_name'])
                ws[f'B{row}'] = clean_string(result['prod_number'])
                ws[f'C{row}'] = result['rating']
                ws[f'D{row}'] = clean_string(result['headline'])
                ws[f'E{row}'] = clean_string(result['review_content'])
                row += 1

        if not os.path.exists(savePath):
            os.mkdir(savePath)

        wb.save(filePath)
        wb.close()

        print(f'파일 저장완료!\n\n{filePath}')